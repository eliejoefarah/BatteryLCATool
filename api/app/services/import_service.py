from __future__ import annotations

# =============================================================================
# import_service.py — Battery LCA Tool
# =============================================================================
# Parses an xlsx file (as bytes) and writes rows to the database.
#
# Supported xlsx formats
# ──────────────────────
# tabular    — 3 flat sheets:
#              • Parameters  (columns: name, description, value, min_value,
#                             max_value, mode_value, distribution_type)
#              • Activities  (columns: name, location, unit, production_amount,
#                             stage, comment)
#              • Exchanges   (columns: activity_name, flow_name, quantity,
#                             formula, unit, direction, source_database,
#                             source_location, data_origin)
#
# brightway  — N sheets, one activity per sheet (Brightway Excel export format
#              as used in example_database.xlsx). Each sheet has:
#              • Key-value metadata rows (Activity, location, unit, …)
#              • An 'Exchanges' row as separator
#              • A table: name | amount | unit | database | location |
#                         type | categories | reference product | comment
#              Exchange types: production → output/reference,
#                              technosphere → input,
#                              biosphere/air → output/waste_output
#
# Format auto-detection:
#   tabular      → 'Activities' and 'Exchanges' sheet names present
#   vub_template → one process per sheet; col A has 'INPUTS' / 'INVENTORY PROCESS'
#   brightway    → one activity per sheet; key-value rows start with 'Activity'
# =============================================================================

import io
import json
import logging
import uuid
from decimal import Decimal, InvalidOperation
from typing import Any, Literal
from uuid import UUID

import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.import_models import (
    BatchImportResult,
    XlsxActivityRow,
    XlsxExchangeRow,
    XlsxParameterRow,
)

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Sheet names that are never process/activity sheets
_SKIP_SHEETS: frozenset[str] = frozenset(
    {
        "version history",
        "readme",
        "read me",
        "example",
        "instructions",
        "legend",
        "changelog",
        "cover",
        "title page",
    }
)

# Brightway exchange type → (direction, output_type)
_BW_TYPE_MAP: dict[str, tuple[str, str | None]] = {
    "production":   ("output", "reference"),
    "technosphere": ("input",  None),
    "biosphere":    ("output", "waste_output"),
    "air":          ("output", "waste_output"),
    "waste":        ("output", "waste_output"),
    "coproduct":    ("output", "coproduct"),
    "substitution": ("output", "coproduct"),
}

_DEFAULT_FLOW_KIND = "material"

# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _to_decimal(v: Any) -> Decimal | None:
    if v is None:
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _hdr(row: tuple[Any, ...]) -> dict[str, int]:
    """Build column-name → column-index map from a header row."""
    return {
        str(c).strip().lower(): i
        for i, c in enumerate(row)
        if c is not None
    }


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------


def _detect_format(wb: openpyxl.Workbook) -> Literal["tabular", "brightway", "vub_template"]:
    lc = {s.lower() for s in wb.sheetnames}
    if "activities" in lc and "exchanges" in lc:
        return "tabular"

    # Sample first non-skip sheet to distinguish brightway vs vub_template
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() in _SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=35, max_col=1, values_only=True):
            first = _str(row[0]) if row else None
            if not first:
                continue
            fl = first.lower()
            # Brightway: starts with 'activity' or 'database' key-value rows
            if fl in ("activity", "database", "format"):
                return "brightway"
            # VUB template: has 'INPUTS' or 'INVENTORY PROCESS' section markers
            if fl in ("inputs", "inventory process") or "process description" in fl:
                return "vub_template"
        break  # only inspect first non-skip sheet

    return "brightway"


# ---------------------------------------------------------------------------
# Tabular format parser
# ---------------------------------------------------------------------------


def _parse_tabular(
    wb_vals: openpyxl.Workbook,
    wb_fmls: openpyxl.Workbook,
) -> tuple[
    list[XlsxActivityRow],
    list[XlsxExchangeRow],
    list[XlsxParameterRow],
    list[str],
]:
    """Parse the 3-sheet flat tabular format. Returns (activities, exchanges, params, errors)."""
    lc_v = {s.lower(): s for s in wb_vals.sheetnames}
    lc_f = {s.lower(): s for s in wb_fmls.sheetnames}

    activities: list[XlsxActivityRow] = []
    exchanges: list[XlsxExchangeRow] = []
    parameters: list[XlsxParameterRow] = []
    errors: list[str] = []

    # ── Parameters ────────────────────────────────────────────────────────
    if "parameters" in lc_v:
        ws = wb_vals[lc_v["parameters"]]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _hdr(rows[0])
            for r_i, row in enumerate(rows[1:], start=2):
                name = _str(row[idx.get("name", 0)])
                val  = _to_decimal(row[idx.get("value", 2)])
                if not name or val is None:
                    continue
                try:
                    parameters.append(XlsxParameterRow(
                        name=name,
                        description=_str(row[idx.get("description", 1)]),
                        value=val,
                        min_value=_to_decimal(row[idx.get("min_value", 3)]),
                        max_value=_to_decimal(row[idx.get("max_value", 4)]),
                        mode_value=_to_decimal(row[idx.get("mode_value", 5)]),
                        distribution_type=_str(row[idx.get("distribution_type", 6)]),
                    ))
                except ValueError as exc:
                    errors.append(f"Parameters row {r_i}: {exc}")

    # ── Activities ────────────────────────────────────────────────────────
    if "activities" in lc_v:
        ws = wb_vals[lc_v["activities"]]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            idx = _hdr(rows[0])
            for r_i, row in enumerate(rows[1:], start=2):
                name = _str(row[idx.get("name", 0)])
                if not name:
                    continue
                prod = _to_decimal(row[idx.get("production_amount", 3)]) or Decimal("1.0")
                try:
                    activities.append(XlsxActivityRow(
                        name=name,
                        location=_str(row[idx.get("location", 1)]),
                        unit=_str(row[idx.get("unit", 2)]),
                        production_amount=prod,
                        stage=_str(row[idx.get("stage", 4)]),
                        comment=_str(row[idx.get("comment", 5)]),
                    ))
                except ValueError as exc:
                    errors.append(f"Activities row {r_i}: {exc}")

    # ── Exchanges ─────────────────────────────────────────────────────────
    if "exchanges" in lc_v:
        ws_v = wb_vals[lc_v["exchanges"]]
        ws_f = wb_fmls[lc_f["exchanges"]] if "exchanges" in lc_f else None
        rows_v = list(ws_v.iter_rows(values_only=True))
        rows_f = list(ws_f.iter_rows(values_only=True)) if ws_f else []
        if rows_v:
            idx = _hdr(rows_v[0])
            qty_col = idx.get("quantity", 2)
            for r_i, row in enumerate(rows_v[1:], start=1):
                act_name  = _str(row[idx.get("activity_name", 0)])
                flow_name = _str(row[idx.get("flow_name", 1)])
                if not act_name or not flow_name:
                    continue

                qty_val = _to_decimal(row[qty_col])

                # Formula: read from formulas workbook (data_only=False)
                formula_str: str | None = None
                if rows_f and r_i < len(rows_f):
                    raw_cell = rows_f[r_i][qty_col] if len(rows_f[r_i]) > qty_col else None
                    if raw_cell and str(raw_cell).startswith("="):
                        formula_str = str(raw_cell)

                direction = _str(row[idx.get("direction", 5)]) or "input"
                try:
                    exchanges.append(XlsxExchangeRow(
                        activity_name=act_name,
                        flow_name=flow_name,
                        quantity=qty_val,
                        formula=formula_str,
                        unit=_str(row[idx.get("unit", 4)]),
                        direction=direction,  # type: ignore[arg-type]
                        source_database=_str(row[idx.get("source_database", 6)]),
                        source_location=_str(row[idx.get("source_location", 7)]),
                        data_origin=_str(row[idx.get("data_origin", 8)]),
                    ))
                except ValueError as exc:
                    errors.append(f"Exchanges row {r_i + 1}: {exc}")

    return activities, exchanges, parameters, errors


# ---------------------------------------------------------------------------
# Brightway format parser
# ---------------------------------------------------------------------------


def _parse_brightway(
    wb_vals: openpyxl.Workbook,
    wb_fmls: openpyxl.Workbook,
) -> tuple[
    list[XlsxActivityRow],
    list[XlsxExchangeRow],
    list[XlsxParameterRow],
    list[str],
]:
    """Parse the Brightway format. Returns (activities, exchanges, params, errors).

    Supports two layouts transparently:
    • One activity per sheet  — classic Brightway Excel export (e.g. example_database.xlsx)
    • Many activities per sheet — activities stacked vertically on one sheet, each block
      starting with a new 'Activity' key-value row (e.g. Ali 2025 SI2 format).

    Within each sheet the parser loops: collect key-value metadata → find 'Exchanges'
    separator → parse exchange table until the next 'Activity' row (or end of sheet).
    For single-activity sheets the inner loop simply runs to the end naturally.
    """
    activities: list[XlsxActivityRow] = []
    exchanges: list[XlsxExchangeRow] = []
    parameters: list[XlsxParameterRow] = []
    errors: list[str] = []

    for sheet_name in wb_vals.sheetnames:
        if sheet_name.lower() in _SKIP_SHEETS:
            continue

        ws_v = wb_vals[sheet_name]
        ws_f = wb_fmls[sheet_name] if sheet_name in wb_fmls.sheetnames else None
        rows_v = list(ws_v.iter_rows(values_only=True))
        rows_f = list(ws_f.iter_rows(values_only=True)) if ws_f else []

        i = 0
        while i < len(rows_v):

            # ── Collect key-value metadata until 'Exchanges' separator ────
            meta: dict[str, Any] = {}
            exchanges_row: int | None = None
            skip_block = False

            while i < len(rows_v):
                row = rows_v[i]
                first = _str(row[0]) if row else None
                if first is None:
                    i += 1
                    continue
                fl = first.lower()
                if fl == "skip":
                    skip_block = True
                    break
                if fl == "exchanges":
                    exchanges_row = i
                    i += 1
                    break
                val = row[1] if len(row) > 1 else None
                meta[fl] = val
                i += 1

            if skip_block:
                break  # skip remainder of this sheet

            # No 'Exchanges' row found — no more activity blocks on this sheet
            if exchanges_row is None or not meta:
                break

            # ── Build activity ────────────────────────────────────────────
            act_name = (
                _str(meta.get("activity"))
                or _str(meta.get("process description:"))
                or sheet_name
            )
            prod_amount = _to_decimal(meta.get("production amount")) or Decimal("1.0")
            try:
                act = XlsxActivityRow(
                    name=act_name,
                    location=_str(meta.get("location")),
                    unit=_str(meta.get("unit")),
                    production_amount=prod_amount,
                    # 'categories' in Brightway maps to our 'stage'
                    stage=_str(meta.get("categories")) or _str(meta.get("stage")),
                    comment=_str(meta.get("comment")),
                )
                activities.append(act)
            except ValueError as exc:
                errors.append(f"Sheet '{sheet_name}' activity '{act_name}': {exc}")
                # Advance to the next 'Activity' row and retry
                while i < len(rows_v):
                    first = _str(rows_v[i][0]) if rows_v[i] else None
                    if first and first.lower() == "activity":
                        break
                    i += 1
                continue

            # ── Skip the exchange-table header row ────────────────────────
            if i >= len(rows_v):
                break
            hdr_row = rows_v[i]
            idx = _hdr(hdr_row)
            amt_col = idx.get("amount", 1)
            i += 1

            # ── Parse exchange rows ───────────────────────────────────────
            # Stop when we hit a new 'Activity' metadata row (multi-activity sheets)
            # or run out of rows (single-activity sheets) — both are handled the same way.
            while i < len(rows_v):
                row = rows_v[i]

                first = _str(row[0]) if row else None
                if first and first.lower() == "activity":
                    break  # start of next activity block; outer loop will collect it

                if not any(v is not None for v in row):
                    i += 1
                    continue

                flow_name = _str(row[idx.get("name", 0)])
                if not flow_name:
                    i += 1
                    continue

                qty_val = _to_decimal(row[amt_col])

                formula_str: str | None = None
                if rows_f and i < len(rows_f) and rows_f[i] is not None:
                    fml_cell = rows_f[i][amt_col] if len(rows_f[i]) > amt_col else None
                    if fml_cell and str(fml_cell).startswith("="):
                        formula_str = str(fml_cell)

                if qty_val is None and formula_str is None:
                    i += 1
                    continue

                bw_type = _str(row[idx.get("type", 5)]) or "technosphere"
                direction, _ = _BW_TYPE_MAP.get(bw_type.lower(), ("input", None))

                try:
                    exchanges.append(XlsxExchangeRow(
                        activity_name=act_name,
                        flow_name=flow_name,
                        quantity=qty_val,
                        formula=formula_str,
                        unit=_str(row[idx.get("unit", 2)]),
                        direction=direction,  # type: ignore[arg-type]
                        source_database=_str(row[idx.get("database", 3)]),
                        source_location=_str(row[idx.get("location", 4)]),
                        data_origin=None,
                    ))
                except ValueError as exc:
                    errors.append(f"Sheet '{sheet_name}' exchange row {i + 1}: {exc}")

                i += 1

    return activities, exchanges, parameters, errors


# ---------------------------------------------------------------------------
# VUB template format parser
# ---------------------------------------------------------------------------


def _parse_vub_template(
    wb_vals: openpyxl.Workbook,
    wb_fmls: openpyxl.Workbook,
) -> tuple[
    list[XlsxActivityRow],
    list[XlsxExchangeRow],
    list[XlsxParameterRow],
    list[str],
]:
    """Parse VUB one-sheet-per-process vertical template format.

    Column mapping (0-indexed in Python):
      Col A (0): Category label / section marker
      Col B (1): Sub-item placeholder — NOT the flow name; ignored for data
      Col C (2): Flow name  ← key field; non-empty = real data row
      Col D (3): Amount
      Col E (4): Unit
      Col F (5): Function/use (inputs) | Treatment/destination (outputs) |
                 Mode of transport (transport)  → comment
      Col G (6): Details / range info                                → details
      Col H (7): Cost (€ per unit)                                   → cost_per_unit
      Col I (8): Origin / location                                   → source_location
      Col J (9): Observations / supplier / recycled-content          → observations
                 (only present in the INPUTS section header)

    Section markers (col A, exact match, case-insensitive, only when col C is empty):
      'inventory process' → metadata preamble; skip header row after it
      'inputs'            → input exchanges follow
      'transport'         → transport inputs (direction='input')
      'outputs'           → output exchanges follow

    Metadata detection (col A contains, case-insensitive):
      'process description' → col B = activity name
      'material/product produced' → col C = ref product, col D = amount, col E = unit

    The reference product is prepended as the first exchange (direction='output')
    so that run_import assigns output_type='reference' to it.
    """
    activities: list[XlsxActivityRow] = []
    exchanges: list[XlsxExchangeRow] = []
    errors: list[str] = []

    for sheet_name in wb_vals.sheetnames:
        if sheet_name.lower() in _SKIP_SHEETS:
            continue

        ws_v = wb_vals[sheet_name]
        ws_f = wb_fmls[sheet_name] if sheet_name in wb_fmls.sheetnames else None
        rows_v = list(ws_v.iter_rows(values_only=True))
        rows_f = list(ws_f.iter_rows(values_only=True)) if ws_f else []

        act_name: str | None = None
        ref_name: str | None = None
        ref_amount: Decimal | None = None
        ref_unit: str | None = None

        state = "meta"           # "meta" | "inputs" | "outputs"
        skip_next_non_empty = False   # True after a section marker — skip the header row

        sheet_exchanges: list[XlsxExchangeRow] = []

        for r_i, row in enumerate(rows_v):
            # Pad row to at least 10 elements to cover col J
            if len(row) < 10:
                row = row + (None,) * (10 - len(row))

            col_a = _str(row[0])
            col_b = _str(row[1])
            col_c = _str(row[2])
            col_d = _to_decimal(row[3])
            col_e = _str(row[4])
            col_f = _str(row[5])   # comment (function/use, treatment, mode of transport)
            col_g = _str(row[6])   # details / range
            col_h = _to_decimal(row[7])  # cost_per_unit
            col_i = _str(row[8])   # source_location (origin)
            col_j = _str(row[9])   # observations

            a_lower = col_a.lower() if col_a else ""

            # ── Section markers ────────────────────────────────────────────
            # Only fire when col C is empty — rows with a flow name in col C
            # are data rows even if col A happens to say "Transport" etc.
            if col_c is None:
                if a_lower == "inventory process":
                    state = "meta"
                    skip_next_non_empty = True
                    continue
                if a_lower == "inputs":
                    state = "inputs"
                    skip_next_non_empty = True
                    continue
                if a_lower == "transport":
                    # Transport items are still inputs
                    state = "inputs"
                    skip_next_non_empty = True
                    continue
                if a_lower == "outputs":
                    state = "outputs"
                    skip_next_non_empty = True
                    continue

            # ── Skip the column header row that follows each section marker ─
            if skip_next_non_empty:
                if col_a is not None or col_b is not None or col_c is not None:
                    skip_next_non_empty = False  # this was the header — stop skipping
                continue  # skip empty rows before the header and the header itself

            # ── Metadata rows ──────────────────────────────────────────────
            if state == "meta":
                if col_a and "process description" in a_lower:
                    act_name = col_b
                elif col_a and "material/product produced" in a_lower:
                    ref_name = col_c
                    ref_amount = col_d
                    ref_unit = col_e
                continue

            # ── Exchange data rows (inputs / outputs) ──────────────────────
            # Only rows where col C is non-empty are real data rows;
            # col B rows (e.g. "Electricity 2") are placeholder labels — ignore them
            if not col_c:
                continue

            # Check for an Excel formula in the amount column (col D, index 3)
            formula_str: str | None = None
            if rows_f and r_i < len(rows_f):
                fml_cell = rows_f[r_i][3] if len(rows_f[r_i]) > 3 else None
                if fml_cell and str(fml_cell).startswith("="):
                    formula_str = str(fml_cell)

            if col_d is None and not formula_str:
                errors.append(
                    f"Sheet '{sheet_name}' row {r_i + 1}: "
                    f"flow '{col_c}' has no quantity or formula; skipped."
                )
                continue

            final_act_name = act_name or sheet_name
            direction: Literal["input", "output"] = (
                "input" if state == "inputs" else "output"
            )

            try:
                sheet_exchanges.append(XlsxExchangeRow(
                    activity_name=final_act_name,
                    flow_name=col_c,
                    quantity=col_d,
                    formula=formula_str,
                    unit=col_e,
                    direction=direction,
                    source_database=None,
                    source_location=col_i,
                    data_origin=None,
                    comment=col_f,
                    details=col_g,
                    cost_per_unit=col_h,
                    observations=col_j,
                ))
            except ValueError as exc:
                errors.append(f"Sheet '{sheet_name}' row {r_i + 1}: {exc}")

        # ── Build activity row ─────────────────────────────────────────────
        final_act_name = act_name or sheet_name
        if not ref_amount:
            errors.append(
                f"Sheet '{sheet_name}': reference product has zero or missing "
                f"production amount; sheet skipped."
            )
            continue
        try:
            act = XlsxActivityRow(
                name=final_act_name,
                location=None,
                unit=ref_unit,
                production_amount=ref_amount,
                stage=None,
                comment=None,
            )
            activities.append(act)
        except ValueError as exc:
            errors.append(f"Sheet '{sheet_name}' activity: {exc}")
            continue

        # ── Prepend reference product as the first exchange ────────────────
        # run_import assigns output_type='reference' to the first output it sees,
        # so inserting the reference product at position 0 is sufficient.
        if ref_name:
            try:
                ref_exc = XlsxExchangeRow(
                    activity_name=final_act_name,
                    flow_name=ref_name,
                    quantity=ref_amount or Decimal("1.0"),
                    formula=None,
                    unit=ref_unit,
                    direction="output",
                    source_database=None,
                    source_location=None,
                    data_origin=None,
                )
                sheet_exchanges.insert(0, ref_exc)
            except ValueError as exc:
                errors.append(f"Sheet '{sheet_name}' reference product: {exc}")

        exchanges.extend(sheet_exchanges)

    # VUB template has no Parameters sheet — return empty list
    return activities, exchanges, [], errors


# ---------------------------------------------------------------------------
# Flow catalog — look up or create
# ---------------------------------------------------------------------------


async def _resolve_flow(
    db: AsyncSession,
    catalog_set_id: UUID,
    canonical_name: str,
    kind: str = _DEFAULT_FLOW_KIND,
    unit: str | None = None,
) -> UUID:
    """Return flow_id for (catalog_set_id, canonical_name, kind), inserting if absent."""
    row = (await db.execute(
        text(
            "SELECT flow_id FROM flow_catalog "
            "WHERE catalog_set_id = :cs AND canonical_name = :name AND kind = :kind"
        ),
        {"cs": str(catalog_set_id), "name": canonical_name, "kind": kind},
    )).fetchone()

    if row:
        return UUID(str(row[0]))

    flow_id = uuid.uuid4()
    await db.execute(
        text("""
            INSERT INTO flow_catalog
              (flow_id, catalog_set_id, canonical_name, display_name, kind, default_unit)
            VALUES (:fid, :cs, :name, :name, :kind, :unit)
            ON CONFLICT (catalog_set_id, canonical_name, kind) DO NOTHING
        """),
        {
            "fid": str(flow_id),
            "cs":  str(catalog_set_id),
            "name": canonical_name,
            "kind": kind,
            "unit": unit,
        },
    )
    # Re-fetch after potential ON CONFLICT DO NOTHING race
    row = (await db.execute(
        text(
            "SELECT flow_id FROM flow_catalog "
            "WHERE catalog_set_id = :cs AND canonical_name = :name AND kind = :kind"
        ),
        {"cs": str(catalog_set_id), "name": canonical_name, "kind": kind},
    )).fetchone()
    return UUID(str(row[0]))


# ---------------------------------------------------------------------------
# Main service entry point
# ---------------------------------------------------------------------------


async def run_import(
    db: AsyncSession,
    content: bytes,
    revision_id: UUID,
    catalog_set_id: UUID,
    imported_by: UUID,
    filename: str,
    force: bool = False,
) -> BatchImportResult:
    """
    Parse *content* (raw xlsx bytes) and persist results to the database.

    Double-import guard
    ───────────────────
    If the revision already has process_instance rows and *force* is False,
    returns immediately with ``already_has_data=True`` and
    ``existing_activities_count`` set. No import_job row is created and
    nothing is written to the DB. The caller should warn the user and
    re-invoke with ``force=True`` to delete the old data and re-import.

    Idempotency
    ───────────
    When *force* is True (or no prior data exists), any existing
    process_instance, process_exchange, and model_parameter rows for
    *revision_id* are deleted before re-inserting. All inserts run inside
    a single transaction; a mid-import error triggers a full rollback.

    Negative quantities
    ───────────────────
    Foreground quantities must be >= 0 (direction carries the sign). A negative
    value logs a warning, uses abs(), and continues — it does NOT abort the import.
    """
    result = BatchImportResult()

    # ── 0. Double-import guard ────────────────────────────────────────────
    existing_row = (await db.execute(
        text("SELECT COUNT(*) FROM process_instance WHERE revision_id = :rid"),
        {"rid": str(revision_id)},
    )).fetchone()
    existing_count = int(existing_row[0]) if existing_row else 0

    if existing_count > 0 and not force:
        result.already_has_data = True
        result.existing_activities_count = existing_count
        return result

    import_job_id = uuid.uuid4()
    result.import_job_id = import_job_id

    # ── 1. Create import_job row immediately (outside main transaction) ───
    await db.execute(
        text("""
            INSERT INTO import_job
              (import_id, revision_id, source_filename, source_format,
               imported_by, status)
            VALUES (:jid, :rid, :fname, 'xlsx', :uid, 'running')
        """),
        {
            "jid":   str(import_job_id),
            "rid":   str(revision_id),
            "fname": filename,
            "uid":   str(imported_by),
        },
    )
    await db.commit()

    try:
        # ── 2. Load valid region codes for location validation ────────────
        region_rows = (await db.execute(text("SELECT code FROM region_catalog"))).fetchall()
        valid_region_codes: set[str] = {r[0] for r in region_rows}

        # ── 3. Open xlsx twice: computed values + raw formulas ────────────
        wb_vals = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        wb_fmls = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=False
        )

        # ── 3. Detect format and parse ────────────────────────────────────
        fmt = _detect_format(wb_vals)
        log.info("Import %s: detected format=%s filename=%s", import_job_id, fmt, filename)

        if fmt == "tabular":
            acts, excs, params, parse_errors = _parse_tabular(wb_vals, wb_fmls)
        elif fmt == "vub_template":
            acts, excs, params, parse_errors = _parse_vub_template(wb_vals, wb_fmls)
        else:
            acts, excs, params, parse_errors = _parse_brightway(wb_vals, wb_fmls)

        wb_vals.close()
        wb_fmls.close()

        for e in parse_errors:
            result.add_warning(e)

        # ── 4. Idempotency — clear previous rows for this revision ────────
        await db.execute(
            text("""
                DELETE FROM process_exchange
                WHERE process_id IN (
                    SELECT process_id FROM process_instance WHERE revision_id = :rid
                )
            """),
            {"rid": str(revision_id)},
        )
        await db.execute(
            text("DELETE FROM process_instance WHERE revision_id = :rid"),
            {"rid": str(revision_id)},
        )
        await db.execute(
            text("DELETE FROM model_parameter WHERE revision_id = :rid"),
            {"rid": str(revision_id)},
        )

        # ── 5. Insert activities → process_instance ───────────────────────
        act_name_to_id: dict[str, UUID] = {}

        for act in acts:
            # Validate location against region_catalog; unknown codes → RoW
            loc = act.location
            if loc is not None and loc not in valid_region_codes:
                result.add_warning(
                    f"Activity '{act.name}': location '{loc}' not found in "
                    f"region_catalog; mapped to 'RoW'."
                )
                loc = "RoW"

            proc_id = uuid.uuid4()
            await db.execute(
                text("""
                    INSERT INTO process_instance
                      (process_id, revision_id, name, location, unit,
                       production_amount, production_unit, stage,
                       system_boundary, comment)
                    VALUES
                      (:pid, :rid, :name, :loc, :unit,
                       :prod, :punit, :stage,
                       'foreground', :comment)
                """),
                {
                    "pid":    str(proc_id),
                    "rid":    str(revision_id),
                    "name":   act.name,
                    "loc":    loc,
                    "unit":   act.unit,
                    "prod":   str(act.production_amount),
                    "punit":  act.unit,   # production_unit mirrors unit from xlsx
                    "stage":  act.stage,
                    "comment": act.comment,
                },
            )
            act_name_to_id[act.name] = proc_id
            result.activities_created += 1

        # ── 6. Insert exchanges → process_exchange ────────────────────────
        # Track which activities have had their first output assigned
        first_output_seen: set[UUID] = set()

        # Cache flow_id lookups within this import to reduce DB round-trips
        flow_cache: dict[tuple[str, str], UUID] = {}  # (canonical_name, kind) → flow_id

        for exc_row in excs:
            proc_id = act_name_to_id.get(exc_row.activity_name)
            if proc_id is None:
                result.add_warning(
                    f"Exchange references unknown activity "
                    f"'{exc_row.activity_name}'; skipped."
                )
                continue

            # Quantity sign validation
            qty = exc_row.quantity
            if qty is not None and qty < 0:
                result.add_warning(
                    f"Negative quantity ({qty}) for flow '{exc_row.flow_name}' "
                    f"in activity '{exc_row.activity_name}'. "
                    f"Foreground quantities must be >= 0 — using abs() value."
                )
                qty = abs(qty)

            # Infer output_type from direction + order
            output_type: str | None = None
            if exc_row.direction == "output":
                if proc_id not in first_output_seen:
                    output_type = "reference"
                    first_output_seen.add(proc_id)
                else:
                    output_type = "waste_output"

            # Infer flow kind
            flow_kind = _DEFAULT_FLOW_KIND
            if exc_row.direction == "output":
                flow_kind = "waste" if output_type != "reference" else "material"

            # Resolve / create flow_catalog entry
            cache_key = (exc_row.flow_name, flow_kind)
            if cache_key not in flow_cache:
                flow_id = await _resolve_flow(
                    db,
                    catalog_set_id,
                    canonical_name=exc_row.flow_name,
                    kind=flow_kind,
                    unit=exc_row.unit,
                )
                flow_cache[cache_key] = flow_id
            else:
                flow_id = flow_cache[cache_key]

            exc_id = uuid.uuid4()
            await db.execute(
                text("""
                    INSERT INTO process_exchange
                      (exchange_id, process_id, flow_id, raw_name,
                       quantity_user, formula_user, user_unit,
                       exchange_direction, output_type,
                       source_database, source_location,
                       comment, details, cost_per_unit, observations,
                       amount_is_ecoinvent_signed)
                    VALUES
                      (:eid, :pid, :fid, :rname,
                       :qty, :fml, :unit,
                       :dir, :otype,
                       :sdb, :sloc,
                       :comment, :details, :cost, :obs,
                       FALSE)
                """),
                {
                    "eid":     str(exc_id),
                    "pid":     str(proc_id),
                    "fid":     str(flow_id),
                    "rname":   exc_row.flow_name,
                    "qty":     str(qty) if qty is not None else None,
                    "fml":     exc_row.formula,
                    "unit":    exc_row.unit,
                    "dir":     exc_row.direction,
                    "otype":   output_type,
                    "sdb":     exc_row.source_database,
                    "sloc":    exc_row.source_location,
                    "comment": exc_row.comment,
                    "details": exc_row.details,
                    "cost":    str(exc_row.cost_per_unit) if exc_row.cost_per_unit is not None else None,
                    "obs":     exc_row.observations,
                },
            )
            result.exchanges_created += 1

        # ── 7. Insert parameters → model_parameter ────────────────────────
        for param in params:
            param_id = uuid.uuid4()
            await db.execute(
                text("""
                    INSERT INTO model_parameter
                      (param_id, revision_id, name, description, value,
                       min_value, max_value, mode_value, distribution_type)
                    VALUES
                      (:pid, :rid, :name, :desc, :val,
                       :minv, :maxv, :modev, :dist)
                    ON CONFLICT (revision_id, name) DO UPDATE
                      SET value             = EXCLUDED.value,
                          description       = EXCLUDED.description,
                          min_value         = EXCLUDED.min_value,
                          max_value         = EXCLUDED.max_value,
                          mode_value        = EXCLUDED.mode_value,
                          distribution_type = EXCLUDED.distribution_type
                """),
                {
                    "pid":   str(param_id),
                    "rid":   str(revision_id),
                    "name":  param.name,
                    "desc":  param.description,
                    "val":   str(param.value),
                    "minv":  str(param.min_value)  if param.min_value  is not None else None,
                    "maxv":  str(param.max_value)  if param.max_value  is not None else None,
                    "modev": str(param.mode_value) if param.mode_value is not None else None,
                    "dist":  param.distribution_type,
                },
            )
            result.parameters_created += 1

        # ── 8. Commit ─────────────────────────────────────────────────────
        await db.commit()

        # ── 9. Update import_job → completed ─────────────────────────────
        await db.execute(
            text("""
                UPDATE import_job
                SET status          = 'completed',
                    activities_count = :acts,
                    exchanges_count  = :excs,
                    warnings_count   = :warns,
                    errors_count     = :errs,
                    log_json         = :log
                WHERE import_id = :jid
            """),
            {
                "acts":  result.activities_created,
                "excs":  result.exchanges_created,
                "warns": len(result.warnings),
                "errs":  len(result.errors),
                "log":   json.dumps({
                    "warnings": result.warnings,
                    "errors":   result.errors,
                }),
                "jid": str(import_job_id),
            },
        )
        await db.commit()

        log.info(
            "Import %s complete: %d activities, %d exchanges, %d parameters, "
            "%d warnings",
            import_job_id,
            result.activities_created,
            result.exchanges_created,
            result.parameters_created,
            len(result.warnings),
        )

    except Exception as exc:
        await db.rollback()
        error_msg = f"Import failed: {exc}"
        result.add_error(error_msg)
        log.exception("Import %s failed for revision %s", import_job_id, revision_id)

        # Best-effort: mark import_job as failed
        try:
            await db.execute(
                text("""
                    UPDATE import_job
                    SET status       = 'failed',
                        errors_count = 1,
                        log_json     = :log
                    WHERE import_id = :jid
                """),
                {
                    "log": json.dumps({"error": error_msg}),
                    "jid": str(import_job_id),
                },
            )
            await db.commit()
        except Exception:
            pass  # DB itself may be broken; don't mask the original error

    return result
