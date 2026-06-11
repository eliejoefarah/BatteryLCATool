from __future__ import annotations

# =============================================================================
# export_service.py — Battery LCA Tool
# =============================================================================
# Generates a VUB-template-compatible .xlsx workbook for a given revision.
#
# Strategy
# ────────
# Loads the bundled import_template.xlsx (which contains a blank "Process 1"
# sheet and a "Version history" sheet with exact styling, merges, and column
# widths), then:
#   1. For each process_instance: copy "Process 1", fill metadata + exchanges.
#   2. Update "Version history" with the export date.
#   3. Append a "Parameters" flat-table sheet.
#   4. Delete the original "Process 1" and "Example"/"Process N" template sheets.
#
# Overflow handling
# ─────────────────
# Each template category has a fixed number of pre-defined slot rows.
# When a category contains more exchanges than slots, extra rows are inserted
# immediately AFTER the last existing slot row for that category (before the
# next category / section header).  All insertions are done in bottom-to-top
# order so earlier row numbers are never disturbed by later insertions.
#
# Unmappable exchanges
# ────────────────────
# Exchanges with mapping_status = 'unmappable' are written normally but their
# data cells (C–J) receive grey fill + strikethrough font.
#
# Public API
# ──────────
#   async def generate_export_xlsx(
#       revision_id:         str,
#       partner_responsible: str,
#       db:                  AsyncSession,
#   ) -> bytes
# =============================================================================

import io
import os
from copy import copy
from datetime import datetime, timezone
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "..", "templates", "import_template.xlsx"
)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

_GREY_FILL = PatternFill(fill_type="solid", fgColor="FFD9D9D9")
_GREY_FONT = Font(color="FF888888", strikethrough=True)
_BOLD_FONT = Font(bold=True)

# ---------------------------------------------------------------------------
# Template category definitions
# ─────────────────────────────
# Each entry:
#   (section, col_a_label, slot_labels, has_ellipsis, first_template_row)
#
# first_template_row = row number in the blank "Process 1" template.
# has_ellipsis = True when the last slot_label is the "…" placeholder (it is
#   kept at the end but never filled with exchange data).
# ---------------------------------------------------------------------------

_INPUT_CATEGORIES = [
    # (section_key, col_a_text, slot_labels, has_ellipsis, first_template_row)
    (
        "electricity",
        "Energy consumption (electricity)",
        ["Electricity 1", "Electricity 2"],
        False,
        24,
    ),
    (
        "heat",
        "Energy consumption (heat)",
        ["Energy 1", "Energy 2"],
        False,
        26,
    ),
    (
        "water",
        "Water consumption",
        ["Water 1"],
        False,
        28,
    ),
    (
        "raw_material",
        "Raw Materials (precursors, gases, solvents, others\u2026)",
        ["Material 1", "Material 2", "Material 3", "Material 4"],
        False,
        29,
    ),
    (
        "component",
        "Components",
        ["Component 1", "Component 2"],
        False,
        33,
    ),
    (
        "other_material",
        "Other materials/substances used within the process (ancillary materials)"
        "                               Including auxilary processes (cleaning,..)",
        ["Material 1", "Material 2", "Material 3", "Material 4"],
        False,
        35,
    ),
    (
        "packaging",
        "Packaging",
        ["Packaging mat. 1", "Packaging mat. 2"],
        False,
        39,
    ),
]

_TRANSPORT_CATEGORY = (
    "transport",
    "Transport",
    ["Transport 1", "Transport 2", "\u2026"],
    True,  # last slot is "…"
    43,
)

_OUTPUT_CATEGORIES = [
    (
        "emission_air",
        "Direct emissions to air",
        ["Emission 1", "Emission 2", "Emission 3", "Emission 4"],
        False,
        48,
    ),
    (
        "emission_water",
        "Emissions to water",
        ["Emission 1", "Emission 2", "Emission 3", "Emission 4"],
        False,
        52,
    ),
    (
        "wastewater",
        "Wastewater produced",
        ["Wwater 1"],
        False,
        56,
    ),
    (
        "solid_waste",
        "Solid waste",
        ["S.waste 1", "S.waste 2"],
        False,
        57,
    ),
    (
        "liquid_waste",
        "Liquid waste",
        ["L.waste1", "\u2026"],
        True,
        59,
    ),
    (
        "other_output",
        "Other outputs (scraps..)",
        ["Scrap 1", "\u2026"],
        True,
        61,
    ),
]

# ---------------------------------------------------------------------------
# Exchange classification
# ---------------------------------------------------------------------------

_TRANSPORT_UNITS = frozenset(
    {"km", "tkm", "t·km", "t km", "tonne-km", "vehicle-km", "ton-km", "pkm", "person-km"}
)
_TRANSPORT_KEYWORDS = frozenset(
    {"transport", "freight", "delivery", "lorry", "truck", "shipping", "haulage", "logistics"}
)


def _classify_input(exc: dict[str, Any]) -> str:
    kind = (exc.get("flow_kind") or "").lower()
    unit = (exc.get("user_unit") or "").lower().strip()
    name = (exc.get("raw_name") or "").lower()

    # Transport check first
    if unit in _TRANSPORT_UNITS or any(kw in name for kw in _TRANSPORT_KEYWORDS):
        return "transport"

    # Electricity vs heat
    if kind == "energy" or unit in {"kwh", "mwh", "gwh", "wh"}:
        if any(x in name for x in ("heat", "steam", "thermal")) or unit in {"mj", "gj", "kj"}:
            return "heat"
        return "electricity"
    if unit in {"mj", "gj", "kj", "j", "kcal", "cal"}:
        return "heat"
    if any(x in name for x in ("heat", "steam", "thermal", "natural gas")):
        return "heat"

    # Water
    if kind == "water" or "water" in name:
        return "water"
    if unit in {"l", "litre", "liter", "liters", "litres", "m3", "m³"} and "water" in name:
        return "water"

    # Packaging
    if "packag" in name:
        return "packaging"

    # Component / infrastructure
    if kind == "infrastructure" or any(x in name for x in ("component", "equipment", "machine", "tool")):
        return "component"

    return "raw_material"


def _classify_output(exc: dict[str, Any]) -> str:
    otype = (exc.get("output_type") or "").lower().replace(" ", "_")
    kind = (exc.get("flow_kind") or "").lower()
    name = (exc.get("raw_name") or "").lower()
    unit = (exc.get("user_unit") or "").lower().strip()

    if otype in ("reference_product", "ref_product", "reference"):
        return "reference"
    if otype in ("co_product", "coproduct", "by_product", "byproduct"):
        return "co_product"

    # Wastewater
    if ("waste" in name and "water" in name) or "wastewater" in name or "sewage" in name:
        return "wastewater"

    # Liquid waste
    if "liquid" in name and "waste" in name:
        return "liquid_waste"
    if kind == "waste" and unit in {"l", "litre", "liter", "m3", "m³"}:
        return "liquid_waste"

    # Emissions to water
    if kind == "emission" and any(x in name for x in ("water", "river", "ocean", "sea", "aqua", "aquatic")):
        return "emission_water"

    # Emissions to air
    if kind == "emission" or any(x in name for x in ("co2", "ghg", "nox", "sox", "pm", "emission", "co ", "ch4", "n2o")):
        return "emission_air"

    # Solid waste
    if kind == "waste" or any(x in name for x in ("solid", "waste", "scrap", "slag", "ash")):
        return "solid_waste"

    return "other_output"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_sheet_name(raw: str, used: set[str]) -> str:
    illegal = r'\/:*?[]'
    cleaned = "".join(c for c in raw if c not in illegal).strip() or "Sheet"
    name = cleaned[:31]
    if name not in used:
        used.add(name)
        return name
    for i in range(2, 1000):
        suffix = f"_{i}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        if candidate not in used:
            used.add(candidate)
            return candidate
    return cleaned[:31]


def _dec_to_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _write_data_cells(
    ws,
    row: int,
    exc: dict[str, Any],
    is_transport: bool = False,
    is_output: bool = False,
) -> None:
    """Write exchange data into columns C–J (or C–I for outputs) of a row."""
    amount = _dec_to_float(exc.get("quantity_user"))
    formula = exc.get("formula_user")
    amount_val = formula if (formula and amount is None) else amount

    unmappable = exc.get("mapping_status") == "unmappable"

    def _cell(col: int, value: Any) -> None:
        c = ws.cell(row=row, column=col, value=value)
        if value is not None and value != "":
            c.font = _GREY_FONT if unmappable else _BOLD_FONT
        if unmappable:
            c.fill = _GREY_FILL

    _cell(3, exc.get("raw_name"))        # C: Name/source
    _cell(4, amount_val)                  # D: Amount / Distance
    _cell(5, exc.get("user_unit"))        # E: Unit
    _cell(6, exc.get("comment"))          # F: Function/use | Mode of transport | Treatment
    _cell(7, exc.get("details"))          # G: Details
    _cell(8, _dec_to_float(exc.get("cost_per_unit")))  # H: Cost

    if not is_output:
        _cell(9, exc.get("source_location"))   # I: Origin
        _cell(10, exc.get("observations"))     # J: Observations
    else:
        # Output section: I+J are merged in template → write to I only
        _cell(9, exc.get("observations"))      # I: Observations


def _set_row_height(ws, row: int, height: float = 15.6) -> None:
    ws.row_dimensions[row].height = height


def _insert_overflow_rows(ws, after_row: int, count: int) -> None:
    """Insert `count` blank rows immediately after `after_row`."""
    if count <= 0:
        return
    ws.insert_rows(after_row + 1, count)
    for r in range(after_row + 1, after_row + 1 + count):
        _set_row_height(ws, r)


# ---------------------------------------------------------------------------
# Process sheet filler
# ---------------------------------------------------------------------------

def _fill_process_sheet(
    ws,
    proc: dict[str, Any],
    exchanges: list[dict[str, Any]],
    partner_responsible: str,
) -> None:
    """Fill data into a copied blank "Process 1" sheet."""

    # ── Classify exchanges ──────────────────────────────────────────────────
    reference_exc: dict | None = None
    co_product_exc: dict | None = None
    inputs: dict[str, list[dict]] = {
        "electricity": [], "heat": [], "water": [],
        "raw_material": [], "component": [], "other_material": [], "packaging": [],
    }
    transport: list[dict] = []
    outputs: dict[str, list[dict]] = {
        "emission_air": [], "emission_water": [], "wastewater": [],
        "solid_waste": [], "liquid_waste": [], "other_output": [],
    }

    for exc in exchanges:
        direction = (exc.get("exchange_direction") or "input").lower()
        if direction in ("output", "reference_product", "co_product", "waste_output"):
            cat = _classify_output(exc)
            if cat == "reference" and reference_exc is None:
                reference_exc = exc
            elif cat == "co_product" and co_product_exc is None:
                co_product_exc = exc
            elif cat in outputs:
                outputs[cat].append(exc)
            else:
                outputs["other_output"].append(exc)
        else:
            cat = _classify_input(exc)
            if cat == "transport":
                transport.append(exc)
            elif cat in inputs:
                inputs[cat].append(exc)
            else:
                inputs["raw_material"].append(exc)

    # ── Static metadata rows ────────────────────────────────────────────────
    # Row 6: process description (name used as primary descriptor)
    ws.cell(row=6, column=2, value=proc.get("name"))
    # Row 7: longer description / comment
    ws.cell(row=7, column=2, value=proc.get("comment"))
    # Row 8: partner responsible
    ws.cell(row=8, column=2, value=partner_responsible)

    # Row 12: reference product
    ref_name = proc.get("name")
    ref_amount = _dec_to_float(proc.get("production_amount"))
    ref_unit = proc.get("unit")
    if reference_exc:
        ref_name = reference_exc.get("raw_name") or ref_name
        ref_amount = _dec_to_float(reference_exc.get("quantity_user")) or ref_amount
        ref_unit = reference_exc.get("user_unit") or ref_unit

    for col, val in [(3, ref_name), (4, ref_amount), (5, ref_unit)]:
        c = ws.cell(row=12, column=col, value=val)
        if val is not None:
            c.font = _BOLD_FONT

    # Row 13: co-product
    if co_product_exc:
        for col, val in [
            (3, co_product_exc.get("raw_name")),
            (4, _dec_to_float(co_product_exc.get("quantity_user"))),
            (5, co_product_exc.get("user_unit")),
        ]:
            c = ws.cell(row=13, column=col, value=val)
            if val is not None:
                c.font = _BOLD_FONT

    # ── Build unified section list and process ALL in bottom-to-top order ───────
    # This is critical: ws.insert_rows() shifts every row BELOW the insertion
    # point. Processing from the bottom-most section upward means each insertion
    # only displaces sections that have already been written, so hardcoded
    # first_template_row values remain accurate for the section currently being
    # processed.

    all_sections: list[tuple] = []

    # Input sections
    for key, a_label, slot_labels, has_ellipsis, t_row in _INPUT_CATEGORIES:
        exc_list = inputs.get(key, [])
        all_sections.append((key, slot_labels, has_ellipsis, t_row, exc_list, False, False))

    # Transport section
    _, _, t_slots, t_ellipsis, t_first_row = _TRANSPORT_CATEGORY
    all_sections.append(("transport", t_slots, t_ellipsis, t_first_row, transport, True, False))

    # Output sections
    for key, a_label, slot_labels, has_ellipsis, t_row in _OUTPUT_CATEGORIES:
        exc_list = outputs.get(key, [])
        all_sections.append((key, slot_labels, has_ellipsis, t_row, exc_list, False, True))

    # Sort descending by first_template_row (bottom section first)
    all_sections.sort(key=lambda s: s[3], reverse=True)

    for key, slot_labels, has_ellipsis, t_row, exc_list, is_transport, is_output in all_sections:
        _fill_section(ws, t_row, slot_labels, has_ellipsis, exc_list,
                      is_transport=is_transport, is_output=is_output)


def _fill_section(
    ws,
    first_row: int,
    slot_labels: list[str],
    has_ellipsis: bool,
    exchanges: list[dict[str, Any]],
    is_transport: bool,
    is_output: bool,
) -> None:
    """
    Fill one category section in the worksheet.

    Slots with data → write exchange data (bold).
    Slots without data → leave as-is (template label only).
    Overflow exchanges → insert new rows, write data.

    The "…" placeholder (if has_ellipsis) is always kept as the last row.
    """
    if has_ellipsis:
        data_slots = slot_labels[:-1]   # numbered slots only
        ellipsis_label = slot_labels[-1]
    else:
        data_slots = slot_labels
        ellipsis_label = None

    n_slots = len(data_slots)
    n_exc = len(exchanges)

    # Insert overflow rows BEFORE the ellipsis (or after the last numbered slot)
    overflow = max(0, n_exc - n_slots)
    if overflow > 0:
        # For sections with ellipsis: insert before the "…" row (= first_row + n_slots)
        # For sections without ellipsis: insert after the last slot (= first_row + n_slots - 1)
        insert_before = first_row + n_slots
        _insert_overflow_rows(ws, insert_before - 1, overflow)
        # insert_before - 1 means "insert after row (insert_before-1)"
        # which is right after the last existing slot row

    # Write data into slots (existing + overflow)
    for i, exc in enumerate(exchanges):
        row = first_row + i
        # Generate overflow label if beyond template slots
        if i >= n_slots:
            # Derive overflow label from the pattern of data_slots[-1]
            overflow_label = _next_slot_label(data_slots, i + 1)
            ws.cell(row=row, column=2, value=overflow_label)
        _write_data_cells(ws, row, exc, is_transport=is_transport, is_output=is_output)

    # Ensure the "…" row is correctly positioned after all data rows
    # (it was shifted by insert_rows automatically — nothing needed)


def _next_slot_label(existing_slots: list[str], n: int) -> str:
    """Generate the n-th slot label extrapolating from the existing pattern."""
    if not existing_slots:
        return str(n)
    last = existing_slots[-1]
    # Try to strip trailing number from last label
    import re
    m = re.match(r'^(.*?)(\d+)$', last)
    if m:
        prefix = m.group(1)
        return f"{prefix}{n}"
    return f"{last} {n}"


# ---------------------------------------------------------------------------
# Version history sheet filler
# ---------------------------------------------------------------------------

def _fill_version_history(
    ws,
    meta: dict[str, Any],
    partner_responsible: str,
) -> None:
    """Add an export row to the Version history sheet (row 3+)."""
    # Find the next empty row (skip header row 1 and template row 2)
    next_row = ws.max_row + 1

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    ws.cell(row=next_row, column=2, value="exported")      # B: Version
    ws.cell(row=next_row, column=3, value=today)           # C: Date
    ws.cell(row=next_row, column=4, value=partner_responsible)  # D: Partner
    ws.cell(row=next_row, column=5, value=meta.get("project_name"))  # E: Author/project
    ws.cell(row=next_row, column=6, value=(                # F: Content
        f"Exported from BatteryLCATool — "
        f"{meta.get('model_name', '')} rev {meta.get('revision_number', '')}"
        + (f" ({meta.get('label')})" if meta.get('label') else "")
    ))


# ---------------------------------------------------------------------------
# Parameters sheet (flat table, no VUB styling required)
# ---------------------------------------------------------------------------

def _add_parameters_sheet(wb, params: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title="Parameters")

    headers = ["Name", "Value", "Distribution", "Min", "Max", "Mode", "Description"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _BOLD_FONT

    for row_i, p in enumerate(params, start=2):
        ws.cell(row=row_i, column=1, value=p.get("name"))
        ws.cell(row=row_i, column=2, value=_dec_to_float(p.get("value")))
        ws.cell(row=row_i, column=3, value=p.get("distribution_type"))
        ws.cell(row=row_i, column=4, value=_dec_to_float(p.get("min_value")))
        ws.cell(row=row_i, column=5, value=_dec_to_float(p.get("max_value")))
        ws.cell(row=row_i, column=6, value=_dec_to_float(p.get("mode_value")))
        ws.cell(row=row_i, column=7, value=p.get("description"))


def _add_mapping_sheet(wb, mapping_rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title="Flow Mapping")

    headers = [
        "Foreground Flow", "Unit", "Direction",
        "Ecoinvent Activity", "Reference Product",
        "Location", "Unit (ecoinvent)",
        "Version", "System Model", "Status",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h).font = _BOLD_FONT

    for row_i, row in enumerate(mapping_rows, start=2):
        ws.cell(row=row_i, column=1, value=row.get("raw_name"))
        ws.cell(row=row_i, column=2, value=row.get("user_unit"))
        ws.cell(row=row_i, column=3, value=row.get("exchange_direction"))
        ws.cell(row=row_i, column=4, value=row.get("confirmed_activity_name"))
        ws.cell(row=row_i, column=5, value=row.get("confirmed_reference_product"))
        ws.cell(row=row_i, column=6, value=row.get("confirmed_location"))
        ws.cell(row=row_i, column=7, value=row.get("confirmed_unit"))
        ws.cell(row=row_i, column=8, value=row.get("confirmed_ecoinvent_version"))
        ws.cell(row=row_i, column=9, value=row.get("confirmed_system_model"))
        ws.cell(row=row_i, column=10, value=row.get("mapping_status"))


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

async def generate_export_xlsx(
    revision_id: str,
    partner_responsible: str,
    db: AsyncSession,
) -> bytes:
    """
    Generate a VUB-template-compatible .xlsx export for the given revision.

    Returns raw .xlsx bytes.
    Raises ValueError if the revision does not exist.
    """

    # ── 1. Fetch revision metadata ────────────────────────────────────────
    meta_row = (await db.execute(
        text("""
            SELECT
                bmr.revision_id,
                bmr.revision_number,
                bmr.label,
                bm.name  AS model_name,
                bm.chemistry,
                bm.functional_unit,
                p.name   AS project_name
            FROM battery_model_revision bmr
            JOIN battery_model bm ON bm.model_id = bmr.model_id
            JOIN project       p  ON p.project_id = bm.project_id
            WHERE bmr.revision_id = :rid
        """),
        {"rid": revision_id},
    )).fetchone()
    if meta_row is None:
        raise ValueError(f"Revision {revision_id!r} not found.")
    meta: dict[str, Any] = dict(meta_row._mapping)

    # ── 2. Fetch processes ────────────────────────────────────────────────
    proc_rows = (await db.execute(
        text("""
            SELECT process_id, name, location, unit, production_amount, stage, comment
            FROM process_instance
            WHERE revision_id = :rid
            ORDER BY created_at
        """),
        {"rid": revision_id},
    )).fetchall()
    processes = [dict(r._mapping) for r in proc_rows]

    # ── 3. Fetch all exchanges ────────────────────────────────────────────
    exc_rows = (await db.execute(
        text("""
            SELECT
                pe.exchange_id, pe.process_id, pe.raw_name,
                pe.quantity_user, pe.formula_user, pe.user_unit,
                pe.exchange_direction, pe.output_type,
                pe.source_database, pe.source_location,
                pe.comment, pe.details, pe.cost_per_unit,
                pe.observations, pe.mapping_status, pe.sort_order,
                fc.kind AS flow_kind
            FROM process_exchange pe
            LEFT JOIN flow_catalog fc ON fc.flow_id = pe.flow_id
            WHERE pe.process_id IN (
                SELECT process_id FROM process_instance WHERE revision_id = :rid
            )
            ORDER BY pe.process_id, pe.sort_order NULLS LAST, pe.created_at
        """),
        {"rid": revision_id},
    )).fetchall()

    exchanges_by_proc: dict[str, list[dict]] = {}
    for exc in [dict(r._mapping) for r in exc_rows]:
        exchanges_by_proc.setdefault(exc["process_id"], []).append(exc)

    # ── 4. Fetch parameters ───────────────────────────────────────────────
    param_rows = (await db.execute(
        text("""
            SELECT name, value, distribution_type, min_value, max_value, mode_value, description
            FROM model_parameter
            WHERE revision_id = :rid
            ORDER BY name
        """),
        {"rid": revision_id},
    )).fetchall()
    params = [dict(r._mapping) for r in param_rows]

    # ── 4b. Fetch mapping data for input exchanges ────────────────────────
    mapping_rows_raw = (await db.execute(
        text("""
            SELECT
                pe.raw_name,
                pe.user_unit,
                pe.exchange_direction,
                bms.confirmed_activity_name,
                bms.confirmed_reference_product,
                bms.confirmed_location,
                bms.confirmed_unit,
                bms.confirmed_ecoinvent_version,
                bms.confirmed_system_model,
                pe.mapping_status
            FROM process_exchange pe
            LEFT JOIN bw_mapping_selection bms ON bms.exchange_id = pe.exchange_id
            JOIN process_instance pi ON pi.process_id = pe.process_id
            WHERE pi.revision_id = :rid
              AND pe.exchange_direction = 'input'
            ORDER BY pe.raw_name
        """),
        {"rid": revision_id},
    )).fetchall()
    mapping_data = [dict(r._mapping) for r in mapping_rows_raw]

    # ── 5. Load and fill workbook ─────────────────────────────────────────
    wb = load_workbook(_TEMPLATE_PATH)

    # Remove sheets we'll recreate or don't need
    for name in ["Example", "Process 2", "Process 3"]:
        if name in wb.sheetnames:
            del wb[name]

    ws_template = wb["Process 1"]
    used_names: set[str] = {"Version history", "Process 1", "Parameters", "Flow Mapping"}

    process_sheets = []
    for proc in processes:
        ws_copy = wb.copy_worksheet(ws_template)
        sheet_name = _safe_sheet_name(proc["name"], used_names)
        ws_copy.title = sheet_name
        _fill_process_sheet(
            ws_copy,
            proc,
            exchanges_by_proc.get(proc["process_id"], []),
            partner_responsible,
        )
        process_sheets.append(ws_copy)

    # Delete the blank template sheet
    del wb["Process 1"]

    # ── 6. Version history ────────────────────────────────────────────────
    if "Version history" in wb.sheetnames:
        _fill_version_history(wb["Version history"], meta, partner_responsible)

    # ── 7. Parameters sheet ───────────────────────────────────────────────
    _add_parameters_sheet(wb, params)

    # ── 7b. Flow Mapping sheet ────────────────────────────────────────────
    _add_mapping_sheet(wb, mapping_data)

    # ── 8. Reorder sheets: process sheets first, then Version history, Parameters, Flow Mapping
    sheet_order = [ws.title for ws in process_sheets] + ["Version history", "Parameters", "Flow Mapping"]
    for i, title in enumerate(sheet_order):
        if title in wb.sheetnames:
            wb.move_sheet(title, offset=wb.sheetnames.index(title) - i)

    # ── 9. Serialise ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
