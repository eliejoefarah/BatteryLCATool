from __future__ import annotations

# =============================================================================
# montecarlo_router.py — Battery LCA Tool
# =============================================================================
# Admin-only Monte Carlo uncertainty analysis endpoints.
#
# Routes (all under /api/v1/montecarlo via __init__.py prefix):
#   POST   /{revision_id}              — trigger a new run (synchronous)
#   GET    /validated-revisions        — list all validated/frozen revisions
#   GET    /{revision_id}/runs         — list past run summaries for a revision
#   GET    /run/{run_id}               — full result for one run
#   DELETE /run/{run_id}               — delete a run record
#   GET    /run/{run_id}/export        — download multi-sheet xlsx
# =============================================================================

import io
import json
import logging
import re
import uuid as _uuid
from datetime import datetime
from typing import Optional
from uuid import UUID

import numpy as np
import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import get_db
from app.routers.bw_mapping_router import get_admin_user_id
from app.services.montecarlo_service import (
    _FlowKey,
    aggregate_statistics,
    compute_sensitivity,
    load_exchanges,
    load_parameters,
    run_montecarlo_loop,
)

log = logging.getLogger(__name__)

router = APIRouter(prefix="/montecarlo", tags=["montecarlo"])

# ---------------------------------------------------------------------------
# Pydantic response models
# ---------------------------------------------------------------------------


class MonteCarloFlowResult(BaseModel):
    flow_name: Optional[str]
    unit: Optional[str]
    direction: str
    mean: float
    std: float
    p5: float
    p25: float
    p50: float
    p75: float
    p95: float
    histogram: dict  # {bin_edges: list[float], counts: list[int]}


class MonteCarloSensitivityRow(BaseModel):
    parameter_name: str
    flow_name: Optional[str]
    spearman_rho: float
    p_value: float


class MonteCarloRunSummary(BaseModel):
    run_id: UUID
    revision_id: UUID
    n_runs: int
    status: str
    created_at: datetime
    completed_at: Optional[datetime] = None
    failed_runs: Optional[int] = None


class MonteCarloRunResponse(MonteCarloRunSummary):
    flows: list[MonteCarloFlowResult]
    sensitivity: list[MonteCarloSensitivityRow]


class ValidatedRevisionRow(BaseModel):
    revision_id: UUID
    revision_name: Optional[str]
    model_name: str
    project_name: str
    status: str
    frozen_at: Optional[datetime] = None
    created_at: datetime
    created_by_email: str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_INVALID_SHEET_CHARS = re.compile(r'[\\/:*?\[\]]')


def _sheet_name(text: str, index: int, max_prefix: int = 26) -> str:
    safe = _INVALID_SHEET_CHARS.sub("_", str(text or ""))
    return f"{safe[:max_prefix]}_{index}"


def _parse_run_row_summary(row: dict) -> MonteCarloRunSummary:
    failed_raw = row.get("failed_runs")
    return MonteCarloRunSummary(
        run_id=row["id"],
        revision_id=row["revision_id"],
        n_runs=row["n_runs"],
        status=row["status"],
        created_at=row["created_at"],
        completed_at=row.get("completed_at"),
        failed_runs=int(failed_raw) if failed_raw is not None else None,
    )


# ---------------------------------------------------------------------------
# POST /{revision_id} — trigger a new Monte Carlo run
# ---------------------------------------------------------------------------


@router.post("/{revision_id}", response_model=MonteCarloRunResponse)
async def trigger_montecarlo(
    revision_id: UUID,
    n_runs: int = Query(default=1000, ge=1, le=10000),
    db: AsyncSession = Depends(get_db),
    admin_id: UUID = Depends(get_admin_user_id),
) -> MonteCarloRunResponse:
    # ── 1. Fetch revision and gate-check status ───────────────────────────
    rev_row = (await db.execute(
        text("SELECT status FROM battery_model_revision WHERE revision_id = :rid"),
        {"rid": str(revision_id)},
    )).fetchone()

    if rev_row is None:
        raise HTTPException(status_code=404, detail="Revision not found.")

    if rev_row[0] not in ("validated", "frozen"):
        raise HTTPException(
            status_code=422,
            detail="Revision must be validated before running Monte Carlo.",
        )

    # ── 2. Pre-flight: check parameters before creating any DB record ─────
    preflight_params = await load_parameters(revision_id, db)

    if not preflight_params:
        raise HTTPException(
            status_code=422,
            detail=(
                "This revision has no model parameters. "
                "Monte Carlo analysis requires at least one parameter with "
                "an uncertainty distribution (normal, lognormal, uniform, or triangular). "
                "Add parameters to the revision before running Monte Carlo."
            ),
        )

    distributional_count = sum(1 for p in preflight_params if p.dist is not None)
    if distributional_count == 0:
        raise HTTPException(
            status_code=422,
            detail=(
                "None of this revision's parameters declare an uncertainty distribution. "
                "At least one parameter must have a distribution set "
                "(normal, lognormal, uniform, or triangular) for Monte Carlo "
                "analysis to produce meaningful results. "
                "Currently all parameters are treated as fixed constants."
            ),
        )

    # ── 3. Insert montecarlo_run with status='running' (pre-flight passed) ─
    run_id = _uuid.uuid4()
    await db.execute(
        text("""
            INSERT INTO montecarlo_run
              (id, revision_id, triggered_by, n_runs, status)
            VALUES (:id, :revision_id, :triggered_by, :n_runs, 'running')
        """),
        {
            "id": str(run_id),
            "revision_id": str(revision_id),
            "triggered_by": str(admin_id),
            "n_runs": n_runs,
        },
    )
    await db.commit()

    # ── 4. Run computation (reuse pre-flight parameters — no second query) ─
    try:
        parameters = preflight_params
        exchange_set = await load_exchanges(revision_id, db)

        run_results, failed_runs, _failed_formulas, samples_successful = (
            run_montecarlo_loop(parameters, exchange_set.elementary_exchanges, n_runs)
        )

        all_flow_keys: list[_FlowKey] = list(
            {key for run in run_results for key in run.keys()}
        )

        flows_stats = aggregate_statistics(run_results, all_flow_keys)

        flow_arrays: dict[_FlowKey, np.ndarray] = {
            key: np.array([run.get(key, 0.0) for run in run_results])
            for key in all_flow_keys
        }
        sensitivity_rows = compute_sensitivity(samples_successful, flow_arrays)

        results_blob = {
            "n_runs": n_runs,
            "failed_runs": failed_runs,
            "flows": flows_stats,
            "sensitivity": sensitivity_rows,
        }

        # ── 4. Update row to completed ────────────────────────────────────
        await db.execute(
            text("""
                UPDATE montecarlo_run
                SET status       = 'completed',
                    completed_at = now(),
                    results      = :results::jsonb
                WHERE id = :run_id
            """),
            {"results": json.dumps(results_blob), "run_id": str(run_id)},
        )
        await db.commit()

        completed_row = (await db.execute(
            text("SELECT created_at, completed_at FROM montecarlo_run WHERE id = :id"),
            {"id": str(run_id)},
        )).fetchone()

        return MonteCarloRunResponse(
            run_id=run_id,
            revision_id=revision_id,
            n_runs=n_runs,
            status="completed",
            created_at=completed_row[0],
            completed_at=completed_row[1],
            failed_runs=failed_runs,
            flows=[MonteCarloFlowResult(**f) for f in flows_stats],
            sensitivity=[MonteCarloSensitivityRow(**s) for s in sensitivity_rows],
        )

    except Exception as exc:
        log.exception("Monte Carlo run %s failed: %s", run_id, exc)
        try:
            await db.rollback()
            await db.execute(
                text("""
                    UPDATE montecarlo_run
                    SET status        = 'failed',
                        error_message = :msg
                    WHERE id = :run_id
                """),
                {"msg": str(exc), "run_id": str(run_id)},
            )
            await db.commit()
        except Exception:
            pass
        raise HTTPException(status_code=422, detail=str(exc)) from exc


# ---------------------------------------------------------------------------
# GET /validated-revisions — list revisions eligible for Monte Carlo
# ---------------------------------------------------------------------------


@router.get("/validated-revisions", response_model=list[ValidatedRevisionRow])
async def list_validated_revisions(
    db: AsyncSession = Depends(get_db),
    _admin_id: UUID = Depends(get_admin_user_id),
) -> list[ValidatedRevisionRow]:
    rows = (await db.execute(
        text("""
            SELECT
                bmr.revision_id,
                bmr.label        AS revision_name,
                bmr.status,
                bmr.frozen_at,
                bmr.created_at,
                bm.name          AS model_name,
                p.name           AS project_name,
                au.email         AS created_by_email
            FROM battery_model_revision bmr
            JOIN battery_model bm ON bm.model_id   = bmr.model_id
            JOIN project        p  ON p.project_id  = bm.project_id
            JOIN app_user       au ON au.user_id    = bmr.created_by
            WHERE bmr.status IN ('validated', 'frozen')
            ORDER BY bmr.created_at DESC
        """),
    )).mappings().all()

    return [
        ValidatedRevisionRow(
            revision_id=row["revision_id"],
            revision_name=row["revision_name"],
            model_name=row["model_name"],
            project_name=row["project_name"],
            status=row["status"],
            frozen_at=row["frozen_at"],
            created_at=row["created_at"],
            created_by_email=row["created_by_email"],
        )
        for row in rows
    ]


# ---------------------------------------------------------------------------
# GET /{revision_id}/runs — list past run summaries for a revision
# ---------------------------------------------------------------------------


@router.get("/{revision_id}/runs", response_model=list[MonteCarloRunSummary])
async def list_runs_for_revision(
    revision_id: UUID,
    db: AsyncSession = Depends(get_db),
    _admin_id: UUID = Depends(get_admin_user_id),
) -> list[MonteCarloRunSummary]:
    rows = (await db.execute(
        text("""
            SELECT
                id,
                revision_id,
                n_runs,
                status,
                created_at,
                completed_at,
                (results->>'failed_runs')::int AS failed_runs
            FROM montecarlo_run
            WHERE revision_id = :rid
            ORDER BY created_at DESC
        """),
        {"rid": str(revision_id)},
    )).mappings().all()

    return [_parse_run_row_summary(dict(row)) for row in rows]


# ---------------------------------------------------------------------------
# GET /run/{run_id} — full result for one run
# ---------------------------------------------------------------------------


@router.get("/run/{run_id}", response_model=MonteCarloRunResponse)
async def get_run(
    run_id: UUID,
    db: AsyncSession = Depends(get_db),
    _admin_id: UUID = Depends(get_admin_user_id),
) -> MonteCarloRunResponse:
    row = (await db.execute(
        text("""
            SELECT id, revision_id, n_runs, status, created_at, completed_at, results
            FROM montecarlo_run
            WHERE id = :run_id
        """),
        {"run_id": str(run_id)},
    )).mappings().fetchone()

    if row is None:
        raise HTTPException(status_code=404, detail="Monte Carlo run not found.")

    results_data = row["results"]
    if isinstance(results_data, str):
        results_data = json.loads(results_data)

    if results_data:
        flows = [MonteCarloFlowResult(**f) for f in results_data.get("flows", [])]
        sensitivity = [
            MonteCarloSensitivityRow(**s) for s in results_data.get("sensitivity", [])
        ]
        failed_runs: Optional[int] = results_data.get("failed_runs")
    else:
        flows = []
        sensitivity = []
        failed_runs = None

    return MonteCarloRunResponse(
        run_id=row["id"],
        revision_id=row["revision_id"],
        n_runs=row["n_runs"],
        status=row["status"],
        created_at=row["created_at"],
        completed_at=row.get("completed_at"),
        failed_runs=failed_runs,
        flows=flows,
        sensitivity=sensitivity,
    )


# ---------------------------------------------------------------------------
# DELETE /run/{run_id} — delete a run record
# ---------------------------------------------------------------------------


@router.delete("/run/{run_id}", status_code=204)
async def delete_run(
    run_id: UUID,
    db: AsyncSession = Depends(get_db),
    _admin_id: UUID = Depends(get_admin_user_id),
) -> None:
    result = await db.execute(
        text("DELETE FROM montecarlo_run WHERE id = :run_id RETURNING id"),
        {"run_id": str(run_id)},
    )
    if result.fetchone() is None:
        raise HTTPException(status_code=404, detail="Monte Carlo run not found.")
    await db.commit()


# ---------------------------------------------------------------------------
# GET /run/{run_id}/export — download multi-sheet xlsx
# ---------------------------------------------------------------------------


@router.get("/run/{run_id}/export")
async def export_run(
    run_id: UUID,
    db: AsyncSession = Depends(get_db),
    _admin_id: UUID = Depends(get_admin_user_id),
) -> StreamingResponse:
    row = (await db.execute(
        text("SELECT results, status FROM montecarlo_run WHERE id = :run_id"),
        {"run_id": str(run_id)},
    )).mappings().fetchone()

    if row is None:
        raise HTTPException(status_code=404, detail="Monte Carlo run not found.")

    if row["status"] != "completed" or row["results"] is None:
        raise HTTPException(
            status_code=422,
            detail="Run is not completed or has no results to export.",
        )

    results_data = row["results"]
    if isinstance(results_data, str):
        results_data = json.loads(results_data)

    flows: list[dict] = results_data.get("flows", [])
    sensitivity: list[dict] = results_data.get("sensitivity", [])
    n_runs: int = results_data.get("n_runs", 0)
    failed_runs: int = results_data.get("failed_runs", 0)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Monte Carlo Run", str(run_id)])
    ws_summary.append(["Total runs", n_runs])
    ws_summary.append(["Failed runs", failed_runs])
    ws_summary.append([])
    ws_summary.append([
        "flow_name", "unit", "direction",
        "mean", "std", "p5", "p25", "p50", "p75", "p95",
    ])
    for f in flows:
        ws_summary.append([
            f.get("flow_name"), f.get("unit"), f.get("direction"),
            f.get("mean"), f.get("std"),
            f.get("p5"), f.get("p25"), f.get("p50"), f.get("p75"), f.get("p95"),
        ])

    # ── Sheets 2..N: per-flow histogram data ─────────────────────────────
    for idx, f in enumerate(flows, start=1):
        title = _sheet_name(f.get("flow_name") or "flow", idx)
        ws = wb.create_sheet(title=title)
        ws.append([
            "flow_name", str(f.get("flow_name")),
        ])
        ws.append(["unit", str(f.get("unit"))])
        ws.append(["direction", str(f.get("direction"))])
        ws.append([])
        hist = f.get("histogram", {})
        bin_edges: list = hist.get("bin_edges", [])
        counts: list = hist.get("counts", [])
        ws.append(["bin_start", "bin_end", "count"])
        for j in range(len(counts)):
            ws.append([
                bin_edges[j] if j < len(bin_edges) else None,
                bin_edges[j + 1] if (j + 1) < len(bin_edges) else None,
                counts[j],
            ])

    # ── Sheet N+1: Sensitivity ────────────────────────────────────────────
    ws_sens = wb.create_sheet(title="Sensitivity")
    ws_sens.append(["parameter_name", "flow_name", "spearman_rho", "p_value"])
    for s in sensitivity:
        ws_sens.append([
            s.get("parameter_name"),
            s.get("flow_name"),
            s.get("spearman_rho"),
            s.get("p_value"),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="montecarlo_{run_id}.xlsx"'
        },
    )
